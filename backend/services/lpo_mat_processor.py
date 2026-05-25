import io
import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.term_parser import parse_term


SHEET_NAME = "LPO MAT - PO"
HEADER_ROW = 6
GROUP_HEADER_ROW = 5
DATA_START_ROW = 7
MONTH_START_YEAR = 2025

JOB_COL = 1
LPO_DATE_COL = 5
DELIVERY_TERMS_COL = 10
PAYMENT_TERMS_COL = 11
LPO_AMOUNT_COL = 18
REMARKS_COL = 24
TOTAL_COL = 25
MONTH_SECTION_START_COL = 27

REQUIRED_HEADERS = {
    JOB_COL: "JOB NUMBER",
    LPO_DATE_COL: "LPO DATE",
    DELIVERY_TERMS_COL: "DELIVERY TERMS",
    PAYMENT_TERMS_COL: "PAYMENT TERMS",
    LPO_AMOUNT_COL: "LPO AMOUNT",
    REMARKS_COL: "REMARKS",
}

PALE_YELLOW = "FFF2CC"
HEADER_BLUE = "D9EAF7"

_THIN = Side(style="thin")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass
class ProcessResult:
    stream: io.BytesIO
    filename: str
    job_numbers_json: str
    summary_json: str


def process_lpo_workbook(
    content: bytes,
    selected_job_numbers_json: str,
    original_filename: str,
) -> ProcessResult:
    selected_job_numbers = _loads_job_numbers(selected_job_numbers_json)
    workbook = load_workbook(io.BytesIO(content))
    values_workbook = load_workbook(io.BytesIO(content), data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain a sheet named '{SHEET_NAME}'.")

    ws = workbook[SHEET_NAME]
    values_ws = values_workbook[SHEET_NAME]

    _validate_headers(ws)
    _ensure_total_header(ws)

    delivery_map, payment_map = _ensure_month_sections(ws)
    summary = _run_task_a(ws, values_ws, delivery_map, payment_map)

    job_numbers = _unique_job_numbers(ws)
    task_b_summary = _run_task_b(ws, values_ws, selected_job_numbers)
    summary.update(task_b_summary)

    _apply_borders(ws)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    stem = Path(original_filename).stem
    filename = f"{stem}-processed.xlsx"

    return ProcessResult(
        stream=output,
        filename=filename,
        job_numbers_json=json.dumps(job_numbers),
        summary_json=json.dumps(summary),
    )


def _loads_job_numbers(raw: str) -> list[str]:
    try:
        value = json.loads(raw or "[]")
    except json.JSONDecodeError:
        return []

    if not isinstance(value, list):
        return []

    return [str(item).strip() for item in value if str(item).strip()]


def _validate_headers(ws):
    for col, expected in REQUIRED_HEADERS.items():
        value = ws.cell(HEADER_ROW, col).value
        normalized = str(value or "").strip().upper()

        if normalized != expected:
            letter = get_column_letter(col)
            raise ValueError(
                f"Expected '{expected}' in {SHEET_NAME}!{letter}{HEADER_ROW}, found '{value}'."
            )


def _ensure_total_header(ws):
    cell = ws.cell(HEADER_ROW, TOTAL_COL)
    cell.value = "Total LPO Amt"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 34


def _month_labels() -> list[str]:
    current_year = datetime.now().year
    labels = []

    for year in range(MONTH_START_YEAR, current_year + 1):
        for month in range(1, 13):
            labels.append(datetime(year, month, 1).strftime("%b-%Y"))

    return labels


def _ensure_month_sections(ws) -> tuple[dict[str, int], dict[str, int]]:
    labels = _month_labels()

    delivery_start = MONTH_SECTION_START_COL
    delivery_end = delivery_start + len(labels) - 1

    payment_start = delivery_end + 1
    payment_end = payment_start + len(labels) - 1

    _replace_merge(ws, GROUP_HEADER_ROW, delivery_start, delivery_end, "DELIVERY")
    _replace_merge(ws, GROUP_HEADER_ROW, payment_start, payment_end, "PAYMENT DUE")

    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)

    for offset, label in enumerate(labels):
        delivery_cell = ws.cell(HEADER_ROW, delivery_start + offset)
        payment_cell = ws.cell(HEADER_ROW, payment_start + offset)

        for cell in [delivery_cell, payment_cell]:
            cell.value = label
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(cell.column)].width = 12

    return (
        {label: delivery_start + i for i, label in enumerate(labels)},
        {label: payment_start + i for i, label in enumerate(labels)},
    )


def _replace_merge(ws, row: int, start_col: int, end_col: int, title: str):
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == row and merged_range.max_row == row:
            overlaps = not (
                merged_range.max_col < start_col or merged_range.min_col > end_col
            )

            if overlaps:
                ws.unmerge_cells(str(merged_range))

    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )

    cell = ws.cell(row, start_col)
    cell.value = title
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)


def _run_task_a(ws, values_ws, delivery_map: dict[str, int], payment_map: dict[str, int]) -> dict:
    total_rows = 0
    closed_rows = 0
    processed_rows = 0
    highlighted_rows = 0
    delivery_cells_written = 0
    payment_cells_written = 0
    highlighted_row_details = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        job_number = str(ws.cell(row, JOB_COL).value or "").strip()
        remarks = str(ws.cell(row, REMARKS_COL).value or "").strip().upper()

        if not job_number:
            continue

        total_rows += 1

        if remarks == "CLOSED":
            closed_rows += 1
            continue

        if remarks != "OPEN":
            continue

        processed_rows += 1

        lpo_date = _as_datetime(ws.cell(row, LPO_DATE_COL).value)
        amount = values_ws.cell(row, LPO_AMOUNT_COL).value

        delivery_term_text = str(ws.cell(row, DELIVERY_TERMS_COL).value or "").strip()
        payment_term_text = str(ws.cell(row, PAYMENT_TERMS_COL).value or "").strip()

        delivery_result = parse_term(ws.cell(row, DELIVERY_TERMS_COL).value, lpo_date)
        payment_result = parse_term(ws.cell(row, PAYMENT_TERMS_COL).value, lpo_date)

        row_unclear = False
        delivery_ok = False
        payment_ok = False

        if delivery_result:
            written = _write_amounts(ws, row, delivery_result.month_labels, delivery_map, amount)
            if written == 0:
                row_unclear = True
            else:
                delivery_cells_written += written
                delivery_ok = True
        else:
            row_unclear = True

        if payment_result:
            written = _write_amounts(ws, row, payment_result.month_labels, payment_map, amount)
            if written == 0:
                row_unclear = True
            else:
                payment_cells_written += written
                payment_ok = True
        else:
            row_unclear = True

        if row_unclear:
            highlighted_rows += 1
            _highlight_row(ws, row)
            highlighted_row_details.append({
                "rowNumber": row,
                "jobNumber": job_number,
                "deliveryTerm": delivery_term_text,
                "paymentTerm": payment_term_text,
                "deliveryOk": delivery_ok,
                "paymentOk": payment_ok,
            })

    delivery_unresolved = sum(1 for d in highlighted_row_details if not d["deliveryOk"])
    payment_unresolved  = sum(1 for d in highlighted_row_details if not d["paymentOk"])

    return {
        "totalRows": total_rows,
        "closedRows": closed_rows,
        "openRows": processed_rows,
        "taskAProcessedOpenRows": processed_rows,
        "taskAHighlightedRows": highlighted_rows,
        "deliveryCellsWritten": delivery_cells_written,
        "paymentCellsWritten": payment_cells_written,
        "highlightedRowDetails": highlighted_row_details,
        "deliveryUnresolvedCount": delivery_unresolved,
        "paymentUnresolvedCount": payment_unresolved,
    }


def _write_amounts(ws, row: int, month_labels: list[str], month_map: dict[str, int], amount) -> int:
    count = 0

    for label in month_labels:
        col = month_map.get(label)

        if col:
            ws.cell(row, col).value = amount
            count += 1

    return count


def _highlight_row(ws, row: int):
    fill = PatternFill("solid", fgColor=PALE_YELLOW)

    for col in range(1, ws.max_column + 1):
        ws.cell(row, col).fill = fill


def _unique_job_numbers(ws) -> list[str]:
    seen = set()
    job_numbers = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        value = str(ws.cell(row, JOB_COL).value or "").strip()

        if value and value not in seen:
            seen.add(value)
            job_numbers.append(value)

    return job_numbers


def _run_task_b(ws, values_ws, selected_job_numbers: list[str]) -> dict:
    totals_written = 0

    if not selected_job_numbers:
        return {"taskBTotalsWritten": 0, "taskBJobTotals": []}

    selected = set(selected_job_numbers)
    totals: dict[str, Decimal] = {}
    last_rows: dict[str, int] = {}
    all_rows: dict[str, list[int]] = {}

    for row in range(DATA_START_ROW, ws.max_row + 1):
        job_number = str(ws.cell(row, JOB_COL).value or "").strip()

        if job_number not in selected:
            continue

        amount = _as_decimal(values_ws.cell(row, LPO_AMOUNT_COL).value)

        if amount is None:
            continue

        totals[job_number] = totals.get(job_number, Decimal("0")) + amount
        last_rows[job_number] = row
        all_rows.setdefault(job_number, []).append(row)

    job_totals = []

    for job_number, total in totals.items():
        row = last_rows[job_number]
        formatted = _format_decimal(total)
        ws.cell(row, TOTAL_COL).value = f"Total LPO_Amount ({job_number}): {formatted}"
        ws.cell(row, TOTAL_COL).fill = PatternFill("solid", fgColor="E2F0D9")
        ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 34
        totals_written += 1
        job_totals.append({
            "jobNumber": job_number,
            "rows": all_rows.get(job_number, []),
            "total": formatted,
            "placedAtRow": row,
        })

    return {"taskBTotalsWritten": totals_written, "taskBJobTotals": job_totals}


def _apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def _as_datetime(value) -> datetime | None:
    if isinstance(value, datetime):
        return value

    if not value:
        return None

    text = str(value).strip()

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    return None


def _as_decimal(value) -> Decimal | None:
    if value is None:
        return None

    if isinstance(value, str) and value.startswith("="):
        return None

    if isinstance(value, float):
        return Decimal(str(round(value, 4)))

    try:
        cleaned = str(value).replace(",", "").strip()
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _format_decimal(value: Decimal) -> str:
    normalized = value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP).normalize()
    return format(normalized, "f")